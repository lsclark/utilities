
"""
Launches searches in a browser based on entries in an XLS(X) file

Assumes a contiguous rectangular range in the file. Joins entries
within a row with ', ', issues a query for each row in the file.
"""

import re
import argparse
import urllib
from pathlib import Path
from openpyxl import load_workbook
import subprocess

_NUMBER = re.compile(r'\D')


def main(args):
    wb = load_workbook(args.xlsx)
    ws = wb.worksheets[args.sheet]
    range = args.range.split(':')
    data = [', '.join(cell.value for cell in row)
            for row in ws[range[0]:range[1]]]
    if args.condition:
        if ':' not in args.condition:
            cols = (args.condition, args.condition)
        else:
            cols = args.condition.split(':')
        crng = (f'{cols[0]}{_NUMBER.sub("",range[0])}',
                f'{cols[1]}{_NUMBER.sub("",range[1])}')
        conds = [any(cell.value for cell in row)
                 for row in ws[crng[0]:crng[1]]]
    else:
        conds = [True for _ in data]

    searches = []
    for search, cond in zip(data, conds):
        if cond:
            search = urllib.parse.quote(search)
            url = args.url.format(search.replace(" ", "+"))
            searches.append(url)

    subprocess.run([args.browser]+searches)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Launch searches in a browser from an XLSX file")
    parser.add_argument("xlsx", help="Path of XLSX file",
                        type=Path)
    parser.add_argument("range",
                        help="Range to parse (in format `start:end` (eg. D5:G9))")
    parser.add_argument("--browser", default="firefox",
                        help="Browser to use (default: firefox)")
    parser.add_argument("--sheet", type=int, default=0,
                        help="Sheet number (default: 0)")
    parser.add_argument("--url",
                        default="https://www.google.com/search?q={}",
                        help="The search pattern as a single-field format-string")
    parser.add_argument("--condition",
                        default=None,
                        help="A column or column range (`C:D`). Only"
                        " searches if one entry of that row is non-empty.")

    args = parser.parse_args()
    main(args)
